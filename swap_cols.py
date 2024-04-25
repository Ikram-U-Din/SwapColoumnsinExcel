import openpyxl

# Load the Excel file
wb = openpyxl.load_workbook('check-odd-col.xlsx')

# Select the active worksheet
ws = wb.active

# Define the total number of columns
total_columns = ws.max_column

# Swap the columns
for col_index in range(1, (total_columns // 2) + 1):
    col_index1 = col_index
    col_index2 = total_columns - col_index + 1
    print(f"Swapping columns {col_index1} and {col_index2}")
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        print(f"Row values before swap: {[cell.value for cell in row]}")
        # Check if both columns exist in the row
        if len(row) >= max(col_index1, col_index2):
            # Perform the swap
            row[col_index1 - 1].value, row[col_index2 - 1].value = row[col_index2 - 1].value, row[col_index1 - 1].value
        print(f"Row values after swap: {[cell.value for cell in row]}")

# Save the changes
wb.save('modified_excel_file.xlsx')
